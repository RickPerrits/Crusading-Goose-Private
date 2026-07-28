from pathlib import Path
from pprint import pformat

from openpyxl import load_workbook


BASE_FOLDER = Path(__file__).parent
EXCEL_FILE = BASE_FOLDER / "monsters_data.xlsx"
OUTPUT_FILE = BASE_FOLDER / "monster_data.py"
SHEET_NAME = "Sheet1"


def clean_text(value):
    """Return trimmed text, or None for blank cells."""
    if value is None:
        return None

    text = str(value).strip()
    return text if text else None


def convert_number(value, field_name, row_number):
    """Convert Excel numeric cells into integers."""
    if value is None or value == "":
        return None

    try:
        return int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(
            f"Row {row_number}: {field_name} must be a number, "
            f"but found {value!r}."
        ) from error


def convert_source_type(value):
    """Convert the spreadsheet copyright marker."""
    text = clean_text(value)

    if text is None:
        return "open"

    normalized = text.lower()

    if normalized in {"yes", "y", "true", "restricted", "copyrighted"}:
        return "restricted"

    if normalized in {"no", "n", "false", "open"}:
        return "open"

    raise ValueError(
        f"Unknown Copywrite value: {value!r}. "
        "Use Yes or No in the spreadsheet."
    )


def import_monsters():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Could not find {EXCEL_FILE.name} in:\n"
            f"{BASE_FOLDER}"
        )

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True,
        read_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Could not find worksheet {SHEET_NAME!r}. "
            f"Available sheets: {workbook.sheetnames}"
        )

    worksheet = workbook[SHEET_NAME]

    headers = {
        cell.value: column_number
        for column_number, cell in enumerate(
            worksheet[1],
            start=1,
        )
        if cell.value is not None
    }

    required_headers = [
        "Monster",
        "Size",
        "Type",
        "Alignment",
        "AC",
        "HP",
        "Threat Level",
        "Damage Dealt",
        "Copywrite",
    ]

    missing_headers = [
        header
        for header in required_headers
        if header not in headers
    ]

    if missing_headers:
        raise ValueError(
            "The spreadsheet is missing these headers: "
            + ", ".join(missing_headers)
        )

    monsters = []
    restricted_count = 0

    for row_number in range(2, worksheet.max_row + 1):
        monster_name = clean_text(
            worksheet.cell(
                row=row_number,
                column=headers["Monster"],
            ).value
        )

        if monster_name is None:
            continue

        threat_level = convert_number(
            worksheet.cell(
                row=row_number,
                column=headers["Threat Level"],
            ).value,
            "Threat Level",
            row_number,
        )

        counter_damage = clean_text(
            worksheet.cell(
                row=row_number,
                column=headers["Damage Dealt"],
            ).value
        )

        if threat_level is None:
            raise ValueError(
                f"Row {row_number}: {monster_name} has no saved "
                "Threat Level value. Open the workbook in Excel, "
                "allow formulas to recalculate, and save it."
            )

        if counter_damage is None:
            raise ValueError(
                f"Row {row_number}: {monster_name} has no saved "
                "Damage Dealt value. Open the workbook in Excel, "
                "allow formulas to recalculate, and save it."
            )

        source_type = convert_source_type(
            worksheet.cell(
                row=row_number,
                column=headers["Copywrite"],
            ).value
        )

        if source_type == "restricted":
            restricted_count += 1

        monster = {
            "name": monster_name,
            "size": clean_text(
                worksheet.cell(
                    row=row_number,
                    column=headers["Size"],
                ).value
            ),
            "creature_type": clean_text(
                worksheet.cell(
                    row=row_number,
                    column=headers["Type"],
                ).value
            ),
            "alignment": clean_text(
                worksheet.cell(
                    row=row_number,
                    column=headers["Alignment"],
                ).value
            ),
            "armor_class": convert_number(
                worksheet.cell(
                    row=row_number,
                    column=headers["AC"],
                ).value,
                "AC",
                row_number,
            ),
            "hit_points": convert_number(
                worksheet.cell(
                    row=row_number,
                    column=headers["HP"],
                ).value,
                "HP",
                row_number,
            ),
            "threat_level": threat_level,
            "counter_damage": counter_damage,
            "source_type": source_type,
        }

        monsters.append(monster)

    workbook.close()

    file_header = '''"""Monster data generated from monsters_data.xlsx.

Do not edit this file manually.
Update the Excel workbook and run import_monsters.py again.
"""

'''

    formatted_monsters = pformat(
        monsters,
        width=100,
        sort_dicts=False,
    )

    OUTPUT_FILE.write_text(
        file_header + f"MONSTERS = {formatted_monsters}\n",
        encoding="utf-8",
    )

    print()
    print("Monster import complete!")
    print(f"Imported monsters: {len(monsters)}")
    print(f"Open monsters: {len(monsters) - restricted_count}")
    print(f"Restricted monsters: {restricted_count}")
    print(f"Output file: {OUTPUT_FILE}")


if __name__ == "__main__":
    try:
        import_monsters()
    except Exception as error:
        print()
        print("Monster import failed:")
        print(error)
        raise